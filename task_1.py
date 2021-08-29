import time
from openpyxl import load_workbook, Workbook
from selenium import webdriver


# from static.util import auth_data, url, expect_title, expect_error, login, password


def test_auth_form(url, is_valid=False, **kwargs):
    browser = webdriver.Chrome()
    for login, password in kwargs.items():
        try:
            browser.get(url)
            lgn = browser.find_element_by_css_selector('input[name="uid"]')
            lgn.clear()
            lgn.send_keys(login)

            pwd = browser.find_element_by_css_selector('input[name="password"]')
            pwd.clear()
            pwd.send_keys(password)

            browser.find_element_by_css_selector('input[name="btnLogin"]').click()

            if is_valid:
                alert = browser.switch_to.alert
                text = alert.text
                alert.accept()
                time.sleep(1)
                if text.strip() in test_data['expect_error']:
                    print('Test case 2: Passed')
                else:
                    print('Test case 2: Failed')
            else:
                curr_title = browser.find_element_by_css_selector('head title').get_attribute('innerHTML')

                if curr_title.strip() in test_data['expect_title']:
                    print('Test case 1: Passed')
                else:
                    print('Test case 1: Failed')

        except Exception as ex:
            print(ex)
    browser.quit()


invalid_data = {}

wb = load_workbook('data.xlsx')
ws = wb.active
for i in range(1, 4):
    invalid_data[ws.cell(row=i, column=1).value] = ws.cell(row=i, column=2).value

login = ws.cell(row=4, column=1).value
password = ws.cell(row=4, column=2).value

test_data = {}

for i in range(5, 8):
    test_data[ws.cell(row=i, column=1).value] = ws.cell(row=i, column=2).value

test_auth_form(url=test_data['url'], is_valid=False, **{login: password})
test_auth_form(url=test_data['url'], is_valid=True, **invalid_data)
